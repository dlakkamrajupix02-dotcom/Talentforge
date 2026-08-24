from __future__ import annotations
from uuid import UUID
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.database import get_db
from app.core.logging import get_logger
from app.models.models import CandidateUser, User
from app.repository.application_repository import get_public_view_jd
from app.schemas.schemas import JobApplicationCreate,JobApplicationResponse,JobApplicationUpdate
from app.services.dependencies import get_current_user, get_current_regular_user
from app.services.application_service import job_application_service


logger = get_logger()
router = APIRouter(prefix="/applications", tags=["Job Applications"])


@router.post("/submit", response_model=JobApplicationResponse, status_code=status.HTTP_201_CREATED)
async def submit_job_application(payload: JobApplicationCreate,db: AsyncSession = Depends(get_db),current_user: User | CandidateUser = Depends(get_current_user)):
    if not current_user.org_id:
        raise HTTPException(status_code=400, detail="User has no organization assigned")
    public_jd = await get_public_view_jd(db, payload.public_jd_id)
    if not public_jd:
        raise HTTPException(status_code=404, detail="Public view JD not found")
    if public_jd.org_id != current_user.org_id:
        raise HTTPException(status_code=403, detail="Access denied to public JD")
    applicant_name = getattr(current_user, "full_name", None) or getattr(current_user, "name", None) or ""
    applicant_email = getattr(current_user, "email", None)
    if not applicant_email:
        raise HTTPException(status_code=400, detail="Applicant email is required")
    original_jd_id = public_jd.parent_jd_id
    if original_jd_id is None:
        raise HTTPException(status_code=400, detail="Public JD does not have an original JD reference")
    try:
        application = await job_application_service.submit_application(db,org_id=current_user.org_id,public_jd_id=payload.public_jd_id,original_jd_id=original_jd_id,
            applicant_name=applicant_name,applicant_email=applicant_email,applicant_phone=None,source=None,metadata=payload.metadata)
        return JobApplicationResponse(id=application.id,org_id=application.org_id,public_jd_id=application.public_jd_id,original_jd_id=application.original_jd_id,
            applicant_name=application.applicant_name,applicant_email=application.applicant_email,applicant_phone=application.applicant_phone,source=application.source,
            status=application.status,interview_stage=application.interview_stage,comments=application.comments,metadata=application.application_metadata,created_at=application.created_at,updated_at=application.updated_at)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logger.exception("Failed to submit job application")
        await db.rollback()
        raise HTTPException(status_code=500, detail="Failed to submit job application")


@router.get("/", response_model=List[JobApplicationResponse])
async def list_job_applications(public_jd_id: Optional[UUID] = Query(None, description="Filter by public JD ID"),status: Optional[str] = Query(None, description="Filter by application status"),
    db: AsyncSession = Depends(get_db),current_user: User | CandidateUser = Depends(get_current_user)):
    if not current_user.org_id:
        raise HTTPException(status_code=400, detail="User has no organization assigned")
    try:
        reviewer_roles = {"Admin", "Super_Admin", "Manager", "HR"}
        if getattr(current_user, "role", None) not in reviewer_roles:
            applicant_email = getattr(current_user, "email", None)
            if not applicant_email:
                raise HTTPException(status_code=400, detail="Current user has no email assigned")
        applications = await job_application_service.list_applications(db,org_id=current_user.org_id,public_jd_id=public_jd_id,status=status,applicant_email=applicant_email)
        return [JobApplicationResponse.model_validate(application) for application in applications]
    except Exception as exc:
        logger.exception("Failed to list job applications")
        raise HTTPException(status_code=500, detail="Failed to list job applications")


@router.get("/export/excel")
async def export_applications_excel(public_jd_id: UUID = Query(..., description="Public JD ID to export applications for"),db: AsyncSession = Depends(get_db),current_user: User = Depends(get_current_regular_user)):
    """Export all applicant details for a given public JD as an Excel (.xlsx) file."""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    if not current_user.org_id:
        raise HTTPException(status_code=400, detail="User has no organization assigned")
    try:
        applications = await job_application_service.list_applications(db, org_id=current_user.org_id, public_jd_id=public_jd_id)
    except Exception as exc:
        logger.exception("Failed to fetch applications for Excel export")
        raise HTTPException(status_code=500, detail="Failed to fetch applications")
    if not applications:
        raise HTTPException(status_code=404, detail="No applications found for this JD")
    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Applicants"
    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin", color="D9D9D9"),right=Side(style="thin", color="D9D9D9"),top=Side(style="thin", color="D9D9D9"),bottom=Side(style="thin", color="D9D9D9"))
    data_font = Font(name="Calibri", size=10)
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    headers = [
        "S.No", "Applicant Name", "Email", "Phone",
        "Status", "Interview Stage", "Source",
        "Comments", "Applied At", "Last Updated",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, app in enumerate(applications, 2):
        values = [
            row_idx - 1,
            app.applicant_name,
            app.applicant_email,
            app.applicant_phone or "",
            app.status or "",
            app.interview_stage or "",
            app.source or "",
            app.comments or "",
            app.created_at.strftime("%Y-%m-%d %H:%M") if app.created_at else "",
            app.updated_at.strftime("%Y-%m-%d %H:%M") if app.updated_at else "",
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-fit column widths
    col_widths = [6, 25, 30, 16, 14, 18, 14, 35, 18, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"applicants_{str(public_jd_id)[:8]}.xlsx"
    return StreamingResponse(buffer,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/{application_id}", response_model=JobApplicationResponse)
async def get_job_application(application_id: UUID, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_regular_user)):
    if not current_user.org_id:
        raise HTTPException(status_code=400, detail="User has no organization assigned")
    application = await job_application_service.get_application(db, application_id, current_user.org_id)
    if not application:
        raise HTTPException(status_code=404, detail="Job application not found")
    return JobApplicationResponse.model_validate(application)


@router.patch("/{application_id}", response_model=JobApplicationResponse)
async def update_job_application(application_id: UUID, payload: JobApplicationUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_regular_user)):
    if not current_user.org_id:
        raise HTTPException(status_code=400, detail="User has no organization assigned")
    update_data = payload.model_dump(exclude_none=True)
    application = await job_application_service.update_application(db, application_id, current_user.org_id, update_data)
    if not application:
        raise HTTPException(status_code=404, detail="Job application not found")
    return JobApplicationResponse.model_validate(application)
